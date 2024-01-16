import snipeitpyapi as snipeit
from snipeitpyapi import Assets
import requests
import json
from importlib import reload
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import date
from deepdiff import DeepDiff
import diff
from ExcelReport import Report
import sys
import logging
from framework_utils import config

###############################
reload(snipeit)

root_path = config.get_root_path()
logger = config.configure_logging(log_file_name="log_py.log", logger_name=__name__)


class Snipe:
    def __init__(self):
        dotenv_path = (root_path + ".env")
        print(dotenv_path)
        load_dotenv(dotenv_path=dotenv_path)
        self.all_assets = snipeit.Assets()
        self.server = os.getenv("server")  # snipe-it server IP
        # self.server = "http://192.168.5.120"
        self.headers = {"accept": "application/json", "Authorization": "Bearer " + os.getenv("token")}
        self.headers_put = {"accept": "application/json", "Authorization": "Bearer " + os.getenv("token"),
                            "content-type": "application/json"}
        # print(self.server)
        self.token = os.getenv("token")  # personal token for snipe API
        self.limit1 = os.getenv("limit1")  # limit for snipe API GET {int} -- None = All
        self.offset1 = os.getenv("offset1")  # offset {int} -- default: {0}
        self.limit2 = os.getenv("limit2")  # limit for snipe API GET {int} -- None = All
        self.offset2 = os.getenv("offset2")  # offset {int} -- default: {0}
        self.total = ""  # number of items assets in snipe {str}

        self.asset_tags = []  # list of all asset tags in snipe-it in order
        self.serial = []  # list of all serial/IMEI in snipe-it in order
        self.supplier = []  # list of suppliers for assets in order
        self.os_number = []  # list of os numbers for assets in order
        self.person = []  # list of people names or usernames for assets in order - if Ready to Deploy then "rtd"
        self.asset_name = []  # list of asset names from snipe-it in order
        self.last_audit_date = []  # list of last audit dates
        self.next_audit_date = []  # list of next audit dates

        self.merged_data = []

        self.export_raw_results_path = (root_path + os.getenv("export_raw_results_path"))
        self.export_pretty_results_path = (root_path + os.getenv("export_pretty_results_path"))
        self.dict_from_snipe_data = {}  # dict wih needed data from snipe-it

        self.total_users = ""
        self.user_dict = {}

        self.asset_dict = {}

        self.accessories_dict = {}
        # headers = {"Accept": "application/json", "Authorization": ("Bearer " + self.token)}
        # response = requests.get(self.server+"/api/v1/hardware", headers=headers)
        # print(response)
        # logger.info(response)

    # append list of all assets from snipe to "merged_data"(!!! MAX - 3000 !!!)
    def get_merged_raw_data_from_snipe(self):
        logger.info("Starting to fetch data from snipeit")
        raw_data_from_snipe = self.all_assets.get(server=self.server, token=self.token, limit=self.limit1,
                                                  offset=self.offset1)
        raw_data_from_snipe_2 = self.all_assets.get(server=self.server, token=self.token, limit=self.limit2,
                                                    offset=self.offset2)
        json_object_snipe = json.loads(raw_data_from_snipe)
        json_object_snipe_2 = json.loads(raw_data_from_snipe_2)

        self.total = json_object_snipe["total"]
        l1_l2 = (json_object_snipe['rows'] + json_object_snipe_2['rows'])
        with open(self.export_raw_results_path + '.raw_data.json', 'w') as outfile:
            json.dump(l1_l2, outfile)
        with open(self.export_raw_results_path + 'raw_data ' + date.today().strftime("%d.%m.%Y") + '.json',
                  'w') as outfile:
            json.dump(l1_l2, outfile)
        with open(self.export_raw_results_path + '.raw_data.json') as j_full:
            self.merged_data = json.load(j_full)
        # return merged_data
        logger.info("Fetched raw data")

    # appends list of all: asset_tags, serials, suppliers, os_numbers - if None = None
    def get_all_data_from_snipe(self):
        logger.info("Starting to process data")
        for i in range(int(self.total)):
            self.asset_tags.append(self.merged_data[i]['asset_tag'])
            self.serial.append(self.merged_data[i]['serial'])

            if self.merged_data[i]['supplier'] is None:
                self.supplier.append(None)
            else:
                self.supplier.append(self.merged_data[i]['supplier']['name'])
            if list(self.merged_data[i]['custom_fields'].keys())[0] == 'ZOPU':
                if self.merged_data[i]['custom_fields']['Broj osnovnog sredstva']['value'] is None:
                    self.os_number.append(None)
                else:
                    self.os_number.append(self.merged_data[i]['custom_fields']['Broj osnovnog sredstva']['value'])
            elif list(self.merged_data[i]['custom_fields'].keys())[0] == 'Broj kartice':
                # print("kartica")
                self.os_number.append(None)
            if self.merged_data[i]["assigned_to"] is None:
                self.person.append("rtd")
            else:
                if self.merged_data[i]["assigned_to"]["name"] is None:
                    self.person.append(self.merged_data[i]["assigned_to"]["username"])
                else:
                    self.person.append(self.merged_data[i]["assigned_to"]["name"])

            if self.merged_data[i]['model']['name'] is None:
                self.asset_name.append("no name")

            else:
                self.asset_name.append(self.merged_data[i]["model"]["name"])

            if self.merged_data[i]['last_audit_date'] is None:
                self.last_audit_date.append(None)
            else:
                self.last_audit_date.append(self.merged_data[i]['last_audit_date']["formatted"])

            if self.merged_data[i]['next_audit_date'] is None:
                self.next_audit_date.append(None)
            else:
                self.next_audit_date.append(self.merged_data[i]['next_audit_date']['formatted'])
        logger.info("Data processing is done")

    # creates dict of needed data from snipe
    def create_dict_from_snipe_data(self):
        logger.info("Starting to create pretty Json")
        keys_for_l2_dict = ["person", "asset_name", "serial", "supplier", "os_number", "last_audit_date",
                            "next_audit_date"]
        self.dict_from_snipe_data = dict.fromkeys(self.asset_tags)
        for key in self.dict_from_snipe_data:
            key_index = self.asset_tags.index(key)
            self.dict_from_snipe_data[key] = dict.fromkeys(keys_for_l2_dict)
            self.dict_from_snipe_data[key]["person"] = self.person[key_index]
            self.dict_from_snipe_data[key]["asset_name"] = self.asset_name[key_index]
            self.dict_from_snipe_data[key]["serial"] = self.serial[key_index]
            self.dict_from_snipe_data[key]["supplier"] = self.supplier[key_index]
            self.dict_from_snipe_data[key]["os_number"] = self.os_number[key_index]
            self.dict_from_snipe_data[key]["last_audit_date"] = self.last_audit_date[key_index]
            self.dict_from_snipe_data[key]["next_audit_date"] = self.next_audit_date[key_index]
        with open(
                self.export_pretty_results_path + 'dict_from_snipe_data ' + date.today().strftime("%d.%m.%Y") + '.json',
                'w') as write_file:
            json.dump(self.dict_from_snipe_data, write_file)
        logger.info("Created pretty Json :)")

    def statement_user_data(self):
        url = f"{self.server}/api/v1/users?limit=300&offset=0&sort=created_at&order=desc&deleted=false&all=false"
        response = requests.get(url, headers=self.headers)
        json_object_snipe = response.json()
        self.total_users = json_object_snipe["total"]
        self.user_dict = {}
        for row in json_object_snipe["rows"]:
            user_data = {k: row[k] for k in ["id", "username", "name", "assets_count"]}
            self.user_dict[row["id"]] = user_data
        return self.user_dict

    def id_from_asset_tag(self, asset_tag):
        json_object_details_from_tag = json.loads(
            self.all_assets.getDetailsByTag(server=self.server, token=self.token, AssetTag=asset_tag))
        id_from_tag = str(json_object_details_from_tag["id"])
        return str(id_from_tag)

    def get_checked_out_assets_by_id(self, user_id):
        url = f"{self.server}/api/v1/users/{user_id}/assets"
        response = requests.get(url, headers=self.headers)
        checked_out_assets = snipeit.Users().getCheckedOutAssets(self.server, self.token, user_id)
        json_data = json.loads(response.text)

        for asset in json_data.get("rows", []):
            asset_id = asset["id"]
            self.asset_dict[asset_id] = {
                "asset_tag": asset["asset_tag"],
                "category": asset["category"]["name"],
                "model": asset["model"]["name"],
                "serial": asset["serial"],
                "card_number": asset["custom_fields"].get("Broj kartice", {}).get("value", ""),
                "card_dec": asset["custom_fields"].get("kartica_decimal - Jantar", {}).get("value", ""),
                "card_hex": asset["custom_fields"].get("kartica_hex- SofaBar", {}).get("value", ""),
                "manufacturer": asset["manufacturer"]["name"]
            }

        return self.asset_dict

    def get_checked_out_accessories_by_id(self, user_id):
        url = f"{self.server}/api/v1/users/{user_id}/accessories"
        response = requests.get(url, headers=self.headers)
        json_data = json.loads(response.text)

        for item in json_data.get("rows", []):
            accessory_id = item["id"]
            category_name = item["category"]["name"]
            manufacturer_name = item["manufacturer"]["name"]
            accessory_name = item["name"]

            self.accessories_dict[accessory_id] = {
                "category_name": category_name,
                "manufacturer_name": manufacturer_name,
                "accessories_name": accessory_name
            }

        return self.accessories_dict

    def update_asset_model_data(self, asset_id, payload):
        url = self.server + "/api/v1/hardware/" + str(asset_id)

        """
        payload = {
            "notes": "null",
            "assigned_to": None,
            "company_id": None,
            "serial": "null",
            "order_number": "null",
            "warranty_months": None,
            "purchase_cost": None,
            "purchase_date": "null",
            "requestable": False,
            "archived": False,
            "rtd_location_id": None,
            "name": "null",
            "location_id": "null"
        }
        """

        response = requests.patch(url, json=payload, headers=self.headers_put)

        print(response.text)

    def get(self):
        self.get_merged_raw_data_from_snipe()
        self.get_all_data_from_snipe()
        self.create_dict_from_snipe_data()


class Update:
    def __init__(self, asset_tag):
        self.asset_tag = asset_tag
        self.snipe = Snipe()

    def set_os_number(self, os_number):
        logger.info(f"Starting to write OS number: {os_number} for asset: {self.asset_tag}")
        device_id = self.snipe.id_from_asset_tag(self.asset_tag)
        payload = f'{{"_snipeit_broj_osnovnog_sredstva_3":"{os_number}"}}'
        self.snipe.all_assets.updateDevice(server=self.snipe.server, token=self.snipe.token, DeviceID=device_id,
                                           payload=payload)
        logger.info(f"Successfully written OS number: {os_number} for asset: {self.asset_tag}")
        return id

    def set_zopu(self):
        logger.info(f"Starting to set ZOPU for asset: {self.asset_tag}")
        device_id = self.snipe.id_from_asset_tag(asset_tag=self.asset_tag)
        payload = '{"_snipeit_zopu_2":"ZOPU"}'
        self.snipe.all_assets.updateDevice(server=self.snipe.server, token=self.snipe.token, DeviceID=device_id,
                                           payload=payload)
        logger.info(f"Successfully set ZOPU for asset: {self.asset_tag}")

    def set_card_dec(self, card_dec):
        logger.info(f"Starting to set card_dec on asset: {self.asset_tag}")
        device_id = self.snipe.id_from_asset_tag(asset_tag=self.asset_tag)
        payload = f'{{"_snipeit_kartica_decimal_jantar_6":"{card_dec}"}}'
        self.snipe.all_assets.updateDevice(server=self.snipe.server, token=self.snipe.token, DeviceID=device_id,
                                           payload=payload)
        logger.info(f"Successfully updated card_dec on asset: {self.asset_tag}")

    def set_card_hex(self, card_hex):
        logger.info(f"Starting to set card_hex on asset: {self.asset_tag}")
        device_id = self.snipe.id_from_asset_tag(asset_tag=self.asset_tag)
        payload = f'{{"_snipeit_kartica_hex_sofabar_5":"{card_hex}"}}'
        self.snipe.all_assets.updateDevice(server=self.snipe.server, token=self.snipe.token, DeviceID=device_id,
                                           payload=payload)
        logger.info(f"Successfully updated card_hex on asset: {self.asset_tag}")


class AccOsData:
    def __init__(self, snipe):
        # root_path = (sys.path[0] + "/")  # /Users/dpustahija1/PycharmProjects/os-snipe_diff/
        self.wb = load_workbook(filename=(root_path + (os.getenv("excel_filename"))))
        self.sheet_ranges = self.wb.active
        self.wb_result = Workbook()
        self.acc_name = []
        self.acc_os_list = []
        self.asset_tags = []
        self.dict_from_acc_data = {}
        self.export_results_path = (root_path + os.getenv("export_results_path_acc"))
        # self.snipe_os_list = snipe.os_number

    # učitava podatke iz tablice u početne liste
    def append_lists_from_excel(self):
        for cell_a in self.sheet_ranges["A"]:
            cell_a_value = cell_a.value
            if cell_a_value == "Inv. broj":
                pass
            else:
                self.acc_os_list.append(str(cell_a_value))

        for cell_b in self.sheet_ranges["B"]:
            cell_b_value = cell_b.value
            if cell_b_value == "Naziv":
                pass
            else:
                self.acc_name.append(str(cell_b_value))

    def get_tags_from_name(self):
        for name in self.acc_name:
            # print(self.acc_name)
            # print(name)
            if name.find("Asset ") != -1:
                tag = (name.split("Asset ", 2)[1])
                if tag[-1] == ".":
                    tag = (tag[:-1])
                else:
                    pass
                if len(tag) == 4:
                    tag = "0" + tag
                self.asset_tags.append(tag)
            else:
                self.asset_tags.append(None)

    def create_dict_from_acc_data(self):
        keys_for_dict = ["asset_name", "os_number", "asset_tag"]
        self.dict_from_acc_data = dict.fromkeys(self.acc_os_list)
        for key in self.dict_from_acc_data:
            key_index = self.acc_os_list.index(key)
            self.dict_from_acc_data[key] = dict.fromkeys(keys_for_dict)
            self.dict_from_acc_data[key]["asset_name"] = self.acc_name[key_index]
            self.dict_from_acc_data[key]["os_number"] = self.acc_os_list[key_index]
            self.dict_from_acc_data[key]["asset_tag"] = self.asset_tags[key_index]
        with open(self.export_results_path + 'dict_from_acc_data ' + date.today().strftime("%d.%m.%Y") + '.json',
                  'w') as write_file:
            json.dump(self.dict_from_acc_data, write_file)

    def get(self):
        self.append_lists_from_excel()
        self.get_tags_from_name()
        self.create_dict_from_acc_data()


class Check:
    def __init__(self, snipe_data={}, acc_os_data={}):
        self.snipe_data = snipe_data
        self.acc_os_data = acc_os_data
        self.matching = []

        self.asset_names_from_snipe_that_match = []
        self.asset_names_from_os_that_match = []
        self.asset_tags_match = []

        self.asset_names_from_os_that_dont_match = []
        self.non_matching = []

        self.rest_tags = []
        self.rest_names = []

    # provjera i return: lista(matching) sa brojevima os koji se nalaze u snipe
    def is_os_in_snipeit(self):
        os_in_snipe_list = []
        for acc_os_num in self.acc_os_data.acc_os_list:
            if acc_os_num in self.snipe_data.os_number:
                # print("ima ga", acc_os_num)
                self.matching.append(acc_os_num)

    # get matching in snipe and os
    def get_matcing(self):
        filename = ("usporedba_match_" + date.today().strftime("%d.%m.%Y"))

        for match in self.matching:
            self.asset_names_from_snipe_that_match.append(
                self.snipe_data.asset_name[int(self.snipe_data.os_number.index(match))])

            self.asset_names_from_os_that_match.append(
                self.acc_os_data.acc_name[int(self.acc_os_data.acc_os_list.index(match))])

            self.asset_tags_match.append(self.snipe_data.asset_tags[int(self.snipe_data.os_number.index(match))])

    # get non-matching from os in snipe
    def get_non_matching(self):
        for non_match in self.acc_os_data.acc_os_list:
            if non_match in self.matching:
                pass
            else:
                # print(non_match)
                self.non_matching.append(non_match)
        for os_n in self.non_matching:
            self.asset_names_from_os_that_dont_match.append(
                self.acc_os_data.acc_name[int(self.acc_os_data.acc_os_list.index(os_n))])

    def get_rest_from_snipe(self):
        # print(self.snipe_data.dict_from_snipe_data, "dict")
        for asset_tag in self.snipe_data.dict_from_snipe_data:
            if (self.snipe_data.dict_from_snipe_data[asset_tag]['os_number'] is None) or (
                    self.snipe_data.dict_from_snipe_data[asset_tag]['os_number'] == ""):
                self.rest_tags.append(asset_tag)
                self.rest_names.append(self.snipe_data.dict_from_snipe_data[asset_tag]['asset_name'])

    def is_asset_tag_valid(self, asset_tag):
        if len(asset_tag) == 4:
            return False
        else:
            return True

    def is_rtd(self, os_number):
        for asset_tag in self.snipe_data.dict_from_snipe_data:
            if self.snipe_data.dict_from_snipe_data[asset_tag]['os_number'] != "":
                if self.snipe_data.dict_from_snipe_data[asset_tag]['os_number'] == os_number:
                    if self.snipe_data.dict_from_snipe_data[asset_tag]['person'] != "rtd":
                        return False
                    else:
                        return True
                        # person = (self.snipe_data.dict_from_snipe_data[asset_tag]['person'])
                        # print(self.snipe_data.dict_from_snipe_data[asset_tag]['os_number'])
                        # asset = {"asset_tag": asset_tag, "person": person, "os_number": os_number}
                        # print(asset)
                        # print(type(asset))
                        # return asset

    def asset_tag_from_os(self, os_number):
        for asset_tag in self.snipe_data.dict_from_snipe_data:
            if self.snipe_data.dict_from_snipe_data[asset_tag]['os_number'] != "":
                if self.snipe_data.dict_from_snipe_data[asset_tag]['os_number'] == os_number:
                    return asset_tag


class Reports:
    def __init__(self):
        # root_path = (sys.path[0] + "/")  # /Users/dpustahija1/PycharmProjects/os-snipe_diff/
        dotenv_path = (root_path + ".env")
        print(dotenv_path)
        load_dotenv(dotenv_path=dotenv_path)
        self.save_path_matching = (os.getenv("export_report_path_matching"))
        self.save_path_non_matching = (os.getenv("export_report_path_non_matching"))
        self.save_path_rest = (os.getenv("export_report_path_rest"))
        self.my_snipe = Snipe()
        self.my_snipe.get()

        my_acc = AccOsData(self.my_snipe)
        my_acc.get()

        self.my_check = Check(snipe_data=self.my_snipe, acc_os_data=my_acc)
        self.my_check.is_os_in_snipeit()
        self.my_check.get_matcing()
        self.my_check.get_non_matching()
        self.my_check.get_rest_from_snipe()

    def matching_snipe_and_os_report(self):
        report = Report(save_path=self.save_path_matching)
        print("starting excel report")
        logger.info("Starting to generate matching snipe and os Excel report")
        report.write_list_to_excel(save_name="matching_snipe_and_os_report", start_column="A", col_name="snipeit name",
                                   lst1=self.my_check.asset_names_from_snipe_that_match)
        report.write_list_to_excel(save_name="matching_snipe_and_os_report", start_column="B", col_name="acc name",
                                   lst1=self.my_check.asset_names_from_os_that_match)
        report.write_list_to_excel(save_name="matching_snipe_and_os_report", start_column="C", col_name="asset tags",
                                   lst1=self.my_check.asset_tags_match)
        report.write_list_to_excel(save_name="matching_snipe_and_os_report", start_column="D", col_name="broj os",
                                   lst1=self.my_check.matching)
        print("ending excel report")
        logger.info("Generated matching snipe and os Excel report")

    def non_matching_snipe_and_os_report(self):
        report = Report(save_path=self.save_path_non_matching)
        print("starting excel report")
        logger.info("Starting to generate non-matching snipe and os Excel report")
        report.write_list_to_excel(save_name="non_matching_snipe_and_os_report", start_column="A", col_name="os broj",
                                   lst1=self.my_check.non_matching)
        report.write_list_to_excel(save_name="non_matching_snipe_and_os_report", start_column="B", col_name="acc name",
                                   lst1=self.my_check.asset_names_from_os_that_dont_match)
        print("ending excel report")
        logger.info("Generated non-matching snipe and os Excel report")

    def rest_in_snipe_report(self):
        report = Report(save_path=self.save_path_rest)
        print("starting excel report")
        logger.info("Starting to generate rest in snipe Excel report")
        report.write_list_to_excel_prefix(save_name="rest_in_snipe_report", start_column="A", col_name="Asset Tag",
                                          prefix_for_tag=True, lst1=self.my_check.rest_tags)
        report.write_list_to_excel(save_name="rest_in_snipe_report", start_column="B", col_name="Snipeit name",
                                   lst1=self.my_check.rest_names)
        print("ending excel report")
        logger.info("Generated rest in snipe Excel report")

    def non_rtd_assets(self, os_numbers):
        asset_list = []
        os_numbers_not_in_snipe = []
        for os_number in os_numbers:
            if os_number in self.my_snipe.os_number:
                if not self.my_check.is_rtd(os_number=os_number):
                    print(os_number)
                    asset_tag = self.my_check.asset_tag_from_os(os_number=os_number)
                    print(asset_tag)
                    asset_json = (json.loads(
                        self.my_snipe.all_assets.getDetailsByTag(server=self.my_snipe.server, token=self.my_snipe.token,
                                                                 AssetTag=asset_tag)))
                    person = asset_json["assigned_to"]["name"]
                    last_checkout = asset_json["last_checkout"].get("formatted", "") if asset_json["last_checkout"] is not None else "Unknown date"
                    asset = {"asset_tag": asset_tag, "person": person, "os_number": os_number, "last_checkout": last_checkout}
                    asset_list.append(asset)
            else:
                os_numbers_not_in_snipe.append(os_number)

        logger.info(f"{os_numbers_not_in_snipe} not in Snipe")
        report = Report(save_path=self.save_path_rest)
        logger.info("Starting to generate non RTD assets Excel report")
        report.write_list_to_excel_prefix(save_name="non_rtd_assets", start_column="A", col_name="Asset Tag",
                                          prefix_for_tag=True, lst1=[item['asset_tag'] for item in asset_list])
        report.write_list_to_excel(save_name="non_rtd_assets", start_column="B", col_name="OS Broj",
                                   lst1=[item['os_number'] for item in asset_list])
        report.write_list_to_excel(save_name="non_rtd_assets", start_column="C", col_name="Odgovorna osoba",
                                   lst1=[item['person'] for item in asset_list])
        report.write_list_to_excel(save_name="non_rtd_assets", start_column="D", col_name="Datum zadnjeg checkout-a",
                                   lst1=[item['last_checkout'] for item in asset_list])
        logger.info("Generated non RTD asset Excel report")
        return asset_list


########################################################################################################################
##########################################################TEST##########################################################
########################################################################################################################
def test():
    my_snipe = Snipe()
    my_snipe.get()

    my_acc = AccOsData(my_snipe)
    my_acc.get()

    my_check = Check(snipe_data=my_snipe, acc_os_data=my_acc)
    my_check.is_os_in_snipeit()
    # my_check.get_matcing()
    # my_check.get_non_matching()
    my_check.get_rest_from_snipe()

    # test_snipe = Snipe()
    # test_snipe.get()

    # print("len: ", len(test_snipe.dict_from_snipe_data))
    # print(test_snipe.total)

    # my_acc = AccOsData(test_snipe)
    # my_acc.get()
    # my_acc.append_lists_from_excel()
    # my_acc.create_dict_from_acc_data()
    # print(my_acc.acc_name)
    # print(my_acc.acc_os_list)
    #  print(my_acc.dict_from_acc_data)
    # ExcelReport.Report().write_list_to_excel(save_name="tst", col_name="name", lst1=test_snipe.person)


def main():
    my_snipe = Snipe()
    # my_snipe.get_merged_raw_data_from_snipe()
    # my_snipe.get_all_data_from_snipe()
    # my_snipe.create_dict_from_snipe_data()
    my_snipe.get()

    # my_acc = AccOsData(my_snipe)
    # my_acc.get()

    # my_check = Check(snipe_data=my_snipe,acc_os_data=my_acc)
    # my_check.is_os_in_snipeit()
    # my_check.get_matcing()
    # my_check.get_non_matching()
    # my_check.get_rest_from_snipe()

    # my_report = ExcelReport()
    # my_report.generate_matching_xlsx(my_check)
    # my_report.generate_non_matching_xlsx(my_check)
    # my_report.generate_rest_xlsx(my_check)


def get_users():
    my_snipe = Snipe()
    # my_snipe.get_checked_out_assets_by_id("4")
    print(my_snipe.get_checked_out_accessories_by_id("155"))

def test_write():
    import cProfile
    import pstats
    with cProfile.Profile() as pr:
        Update().os_number_list(asset_tag="1", os_number="1001")
    stats = pstats.Stats(pr)
    stats.sort_stats(pstats.SortKey.TIME)
    stats.print_stats()


def my_diff():
    # diff.Diff("dict_from_snipe_data_10.10.2022", "dict_from_snipe_data 11.10.2022", save_name="t", save_path="").pretty_diffs_xlsx()
    diff.Diff("dict_from_snipe_data_10.10.2022", "dict_from_snipe_data_11.10.2022", save_name="g",
              save_path="results_cron/diff/").pretty_diffs_xlsx()


def test_is_rtd():
    my_report = Reports()
    list = my_report.non_rtd_assets(os_numbers=["129", "508", "1222"])
    print(list)


if __name__ == "__main__":
    # test_is_rtd()
    # my_diff()
    # main()
    get_users()
    # test()
    # Reports().matching_snipe_and_os_report()
    # Reports().non_matching_snipe_and_os_report()
    # Reports().rest_in_snipe_report()
    # test_write()
