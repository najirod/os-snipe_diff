import snipeit
from snipeit import Assets
import json
from importlib import reload
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import date
from jsondiff import diff
###############################
reload(snipeit)
load_dotenv()
###############################
server = os.getenv("server")  # snipe-it server IP
token = os.getenv("token")  # personal token for snipe API
limit = os.getenv("limit")  # limit for snipe API GET {int} -- None = All -- MAX je 500!!!
offset = os.getenv("offset")  # offset {int} -- default: {0}
total = ""      # number of items assets in snipe {str}
merged_data = []    # list of dicts with all data in snipe-it
asset_tags = []   # list of all asset tags in snipe-it in order
serial = []     # list of all serial/IMEI in snipe-it in order
supplier = []   # list of suppliers for assets in order
os_number = []  # list of os numbers for assets in order
person = []     # list of people names or usernames for assets in order - if Ready to Deploy then "rtd"
asset_name = []     # list of asset names from snipe-it in order
dict_from_snipe_data = {}   # dict wih needed data from snipe-it

matching = []   # list of matching os numbers in snipe-it & accounting tables
non_matching = []   # list of non-existing os number in snipe-it from accounting tables

export_results_path = os.getenv("export_results_path")

wb = load_workbook(filename=os.getenv("excel_filename"))
sheet_ranges = wb["1"]
wb_result = Workbook()
all_assets = snipeit.Assets()

acc_name = []
acc_os_list = []
snipe_os_list = os_number
###############################


# append list of all assets from snipe to "merged_data"(!!! MAX - 1998 !!!)
def merged_raw_data_from_snipe():
    global total
    global merged_data
    merged_data = []
    raw_data_from_snipe = all_assets.get(server=server, token=token, limit=limit, offset=offset)
    raw_data_from_snipe_2 = all_assets.get(server=server, token=token, limit=500, offset=500)
    raw_data_from_snipe_3 = all_assets.get(server=server, token=token, limit=499, offset=1000)
    raw_data_from_snipe_4 = all_assets.get(server=server, token=token, limit=499, offset=1499)
    json_object_snipe = json.loads(raw_data_from_snipe)
    json_object_snipe_2 = json.loads(raw_data_from_snipe_2)
    json_object_snipe_3 = json.loads(raw_data_from_snipe_3)
    json_object_snipe_4 = json.loads(raw_data_from_snipe_4)

    total = json_object_snipe["total"]
    l1_l2 = (json_object_snipe['rows'] + json_object_snipe_2['rows'] + json_object_snipe_3['rows'] +
             json_object_snipe_4['rows'])
    with open(export_results_path + 'raw_data.json', 'w') as outfile:
        json.dump(l1_l2, outfile)
    with open(export_results_path + 'raw_data.json') as j_full:
        merged_data = json.load(j_full)
    return merged_data


# appends list of all: asset_tags, serials, suppliers, os_numbers - if None = ""
def get_all_data_from_snipe():
    global asset_tags
    global serial
    global supplier
    global os_number
    global person
    global asset_name

    for i in range(int(total)):
        asset_tags.append(merged_data[i]['asset_tag'])
        serial.append(merged_data[i]['serial'])

        if merged_data[i]['supplier'] is None:
            supplier.append(None)
        else:
            supplier.append(merged_data[i]['supplier']['name'])

        if merged_data[i]['custom_fields']['Broj osnovnog sredstva']['value'] is None:
            os_number.append(None)
        else:
            os_number.append(merged_data[i]['custom_fields']['Broj osnovnog sredstva']['value'])
        if merged_data[i]["assigned_to"] is None:
            person.append("rtd")
        else:
            if merged_data[i]["assigned_to"]["name"] is None:
                person.append(merged_data[i]["assigned_to"]["username"])
            else:
                person.append(merged_data[i]["assigned_to"]["name"])

        if merged_data[i]['model']['name'] is None:
            asset_name.append("no name")

        else:
            asset_name.append(merged_data[i]["model"]["name"])


# creates dict of needed data from snipe
def create_dict_from_snipe_data():
    keys_for_l2_dict = ["person","asset_name","serial","supplier","os_number"]
    global dict_from_snipe_data
    dict_from_snipe_data = dict.fromkeys(asset_tags)
    for key in dict_from_snipe_data:
        key_index = asset_tags.index(key)
        dict_from_snipe_data[key]=dict.fromkeys(keys_for_l2_dict)
        dict_from_snipe_data[key]["person"] = person[key_index]
        dict_from_snipe_data[key]["asset_name"] = asset_name[key_index]
        dict_from_snipe_data[key]["serial"] = serial[key_index]
        dict_from_snipe_data[key]["supplier"] = supplier[key_index]
        dict_from_snipe_data[key]["os_number"] = os_number[key_index]
    with open(export_results_path + 'dict_from_snipe_data.json','w') as write_file:
        json.dump(dict_from_snipe_data,write_file)

# učitava podatke iz tablice u početne liste
def append_lists_from_excel():
    """
    for cell_a in sheet_ranges["A"]:
        cell_a_value = isinstance(cell_a.value, int)
        if cell_a_value:
            if cell_a.value not in acc_os_list:
                acc_os_list.append(str(cell_a.value))
    """
    for cell_a in sheet_ranges["A"]:
        cell_a_value = cell_a.value
        if cell_a_value == "Inv. broj":
            pass
        else:
            acc_os_list.append(str(cell_a_value))

    for cell_b in sheet_ranges["B"]:
        cell_b_value = cell_b.value
        if cell_b_value == "Naziv":
            pass
        else:
            acc_name.append(str(cell_b_value))
    """"
    for cell_c in sheet_ranges["C"]:
        cell_c_value = isinstance(cell_c.value, int)
        if cell_c_value:
            if cell_c.value not in snipe_os_list:
                snipe_os_list.append(cell_c.value)
    """


# provjera i return: lista(matching) sa brojevima os koji se nalaze u snipe
def is_os_in_snipeit():
    global matching
    os_in_snipe_list = []
    for acc_os_num in acc_os_list:
        if acc_os_num in snipe_os_list:
            # print("ima ga", acc_os_num)
            matching.append(acc_os_num)


# upis podataka u excel
def write_to_excel(save_name="result", start_column="A", os_from="snipe or acc", lst1=None, lst2=None):
    sheet_ranges_result = wb_result.active
    cell_n = 2
    print(cell_n)
    if lst2 is None:
        for acc_os in lst1:
            if acc_os is None:
                pass
            else:
                sheet_ranges_result[start_column + "1"] = "os_from " + os_from
                sheet_ranges_result[start_column + str(cell_n)] = acc_os
                cell_n += 1
                wb_result.save(export_results_path + save_name+ ".xlsx")
    else:
        for acc_os in lst1:
            if acc_os is None:
                pass
            else:
                sheet_ranges_result[start_column+"1"] = "os_from " + os_from
                sheet_ranges_result[start_column + str(cell_n)] = acc_os
                sheet_ranges_result[chr(ord(start_column) + 1) + "1"] = "name"
                sheet_ranges_result[chr(ord(start_column)+1)+str(cell_n)] = lst2[cell_n]
                cell_n += 1
                wb_result.save(export_results_path + save_name + ".xlsx")
    print(cell_n)


# get matching in snipe and os & create xlsx
def get_matcing():
    aseet_names_from_snipe_that_match = []
    asset_names_from_os_that_match = []
    asset_tags_match = []
    filename = ("usporedba_match_" + date.today().strftime("%d.%m.%Y"))
    for match in matching:
        #print(match)
        # print(matching.index(match))

        # print(snipe_os_list.index(match))
       # print(snipe_os_list[snipe_os_list.index(match)])

        aseet_names_from_snipe_that_match.append(asset_name[int(snipe_os_list.index(match))])
        # print(asset_name[int(snipe_os_list.index(match))])

        asset_names_from_os_that_match.append(acc_name[int(acc_os_list.index(match))])
        # print(acc_name[int(acc_os_list.index(match))])
        # print(matching)

        asset_tags_match.append(asset_tags[int(snipe_os_list.index(match))])
    print(len(aseet_names_from_snipe_that_match))
    print(len(asset_names_from_os_that_match))
    print(len(asset_tags_match))
    write_to_excel(save_name=filename, start_column="A", os_from="nums", lst1=matching)
    write_to_excel(save_name=filename, start_column="B", os_from="name_snipe", lst1=aseet_names_from_snipe_that_match)
    write_to_excel(save_name=filename, start_column="D", os_from="name_acc", lst1=asset_names_from_os_that_match)
    write_to_excel(save_name=filename, start_column="C", os_from="asset_tag", lst1=asset_tags_match)


# get non-matching from os in snipe & create xlsx
def get_non_matching():
    asset_names_from_os_that_dont_match = []
    filename = ("not in snipe "+ date.today().strftime("%d.%m.%Y"))
    global non_matching
    for non_match in acc_os_list:
        if non_match in matching:
            pass
        else:
            #print(non_match)
            non_matching.append(non_match)
    for os_n in non_matching:
        asset_names_from_os_that_dont_match.append(acc_name[int(acc_os_list.index(os_n))])

    write_to_excel(save_name=filename, start_column="A",os_from="acc",lst1=non_matching)
    write_to_excel(save_name=filename, start_column="B",os_from="nam_in_os",lst1=asset_names_from_os_that_dont_match)


def main():
    global matching
    merged_raw_data_from_snipe()
    get_all_data_from_snipe()
    append_lists_from_excel()
    is_os_in_snipeit()
    create_dict_from_snipe_data()


def optional():
    #get_matcing()
    get_non_matching()
    #test


def test_diff():

    # diff(json1, json2)
    pass


if __name__ == "__main__":
    main()
    optional()

