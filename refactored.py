import snipeit
from snipeit import Assets
import json
from importlib import reload
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import date
from deepdiff import DeepDiff
import diff
###############################
reload(snipeit)
# load_dotenv()
###############################
# server = os.getenv("server")  # snipe-it server IP
# token = os.getenv("token")  # personal token for snipe API
# limit = os.getenv("limit")  # limit for snipe API GET {int} -- None = All -- MAX je 500!!!
# offset = os.getenv("offset")  # offset {int} -- default: {0}
# total = ""      # number of items assets in snipe {str}
# merged_data = []    # list of dicts with all data in snipe-it
# asset_tags = []   # list of all asset tags in snipe-it in order
# serial = []     # list of all serial/IMEI in snipe-it in order
# supplier = []   # list of suppliers for assets in order
# os_number = []  # list of os numbers for assets in order
# person = []     # list of people names or usernames for assets in order - if Ready to Deploy then "rtd"
# asset_name = []     # list of asset names from snipe-it in order
# dict_from_snipe_data = {}   # dict wih needed data from snipe-it

# matching = []   # list of matching os numbers in snipe-it & accounting tables
# non_matching = []   # list of non-existing os number in snipe-it from accounting tables


# export_results_path = os.getenv("export_results_path")

"""
wb = load_workbook(filename=os.getenv("excel_filename"))
sheet_ranges = wb["1"]
wb_result = Workbook()
"""
# all_assets = snipeit.Assets()

# acc_name = []
# acc_os_list = []
# # snipe_os_list = os_number
###############################


class Snipe:
    def __init__(self):
        load_dotenv()
        self.all_assets = snipeit.Assets()
        self.server = os.getenv("server")  # snipe-it server IP
        self.token = os.getenv("token")  # personal token for snipe API
        self.limit = os.getenv("limit")  # limit for snipe API GET {int} -- None = All -- MAX je 500!!!
        self.offset = os.getenv("offset")  # offset {int} -- default: {0}
        self.total = ""  # number of items assets in snipe {str}

        self.asset_tags = []  # list of all asset tags in snipe-it in order
        self.serial = []  # list of all serial/IMEI in snipe-it in order
        self.supplier = []  # list of suppliers for assets in order
        self.os_number = []  # list of os numbers for assets in order
        self.person = []  # list of people names or usernames for assets in order - if Ready to Deploy then "rtd"
        self.asset_name = []  # list of asset names from snipe-it in order

        self.merged_data = []

        self.export_results_path = os.getenv("export_results_path_test")
        self.dict_from_snipe_data = {}  # dict wih needed data from snipe-it

    # append list of all assets from snipe to "merged_data"(!!! MAX - 1998 !!!)
    def get_merged_raw_data_from_snipe(self):
        raw_data_from_snipe = self.all_assets.get(server=self.server, token=self.token, limit=self.limit, offset=self.offset)
        raw_data_from_snipe_2 = self.all_assets.get(server=self.server, token=self.token, limit=500, offset=500)
        raw_data_from_snipe_3 = self.all_assets.get(server=self.server, token=self.token, limit=499, offset=1000)
        raw_data_from_snipe_4 = self.all_assets.get(server=self.server, token=self.token, limit=499, offset=1499)
        json_object_snipe = json.loads(raw_data_from_snipe)
        json_object_snipe_2 = json.loads(raw_data_from_snipe_2)
        json_object_snipe_3 = json.loads(raw_data_from_snipe_3)
        json_object_snipe_4 = json.loads(raw_data_from_snipe_4)

        self.total = json_object_snipe["total"]
        l1_l2 = (json_object_snipe['rows'] + json_object_snipe_2['rows'] + json_object_snipe_3['rows'] +
                 json_object_snipe_4['rows'])
        with open(self.export_results_path + 'raw_data.json', 'w') as outfile:
            json.dump(l1_l2, outfile)
        with open(self.export_results_path + 'raw_data.json') as j_full:
            self.merged_data = json.load(j_full)
        # return merged_data

    # appends list of all: asset_tags, serials, suppliers, os_numbers - if None = None
    def get_all_data_from_snipe(self):

        for i in range(int(self.total)):
            self.asset_tags.append(self.merged_data[i]['asset_tag'])
            self.serial.append(self.merged_data[i]['serial'])

            if self.merged_data[i]['supplier'] is None:
                self.supplier.append(None)
            else:
                self.supplier.append(self.merged_data[i]['supplier']['name'])
            if list(self.merged_data[i]['custom_fields'].keys())[0] == 'ZOPU':
                if self.merged_data[i]['custom_fields']['Broj osnovnog sredstva']['value'] is None:
                    self.os_number.append(None)
                else:
                    self.os_number.append(self.merged_data[i]['custom_fields']['Broj osnovnog sredstva']['value'])
            elif list(self.merged_data[i]['custom_fields'].keys())[0] == 'Broj kartice':
                # print("kartica")
                self.os_number.append(None)
            if self.merged_data[i]["assigned_to"] is None:
                self.person.append("rtd")
            else:
                if self.merged_data[i]["assigned_to"]["name"] is None:
                    self.person.append(self.merged_data[i]["assigned_to"]["username"])
                else:
                    self.person.append(self.merged_data[i]["assigned_to"]["name"])

            if self.merged_data[i]['model']['name'] is None:
                self.asset_name.append("no name")

            else:
                self.asset_name.append(self.merged_data[i]["model"]["name"])

    # creates dict of needed data from snipe
    def create_dict_from_snipe_data(self):
        keys_for_l2_dict = ["person", "asset_name", "serial", "supplier", "os_number"]
        self.dict_from_snipe_data = dict.fromkeys(self.asset_tags)
        for key in self.dict_from_snipe_data:
            key_index = self.asset_tags.index(key)
            self.dict_from_snipe_data[key] = dict.fromkeys(keys_for_l2_dict)
            self.dict_from_snipe_data[key]["person"] = self.person[key_index]
            self.dict_from_snipe_data[key]["asset_name"] = self.asset_name[key_index]
            self.dict_from_snipe_data[key]["serial"] = self.serial[key_index]
            self.dict_from_snipe_data[key]["supplier"] = self.supplier[key_index]
            self.dict_from_snipe_data[key]["os_number"] = self.os_number[key_index]
        with open(self.export_results_path + 'dict_from_snipe_data.json', 'w') as write_file:
            json.dump(self.dict_from_snipe_data, write_file)

class write_to():
    """TODO: write to Snipe"""

class AccOsData:
    def __init__(self, snipe):
        self.wb = load_workbook(filename=os.getenv("excel_filename"))
        self.sheet_ranges = self.wb["1"]
        self.wb_result = Workbook()
        self.acc_name = []
        self.acc_os_list = []
        # self.snipe_os_list = snipe.os_number

    # učitava podatke iz tablice u početne liste
    def append_lists_from_excel(self):
        for cell_a in self.sheet_ranges["A"]:
            cell_a_value = cell_a.value
            if cell_a_value == "Inv. broj":
                pass
            else:
                self.acc_os_list.append(str(cell_a_value))

        for cell_b in self.sheet_ranges["B"]:
            cell_b_value = cell_b.value
            if cell_b_value == "Naziv":
                pass
            else:
                self.acc_name.append(str(cell_b_value))

class Check:
    def __init__(self, snipe_data, acc_os_data):
        self.snipe_data = snipe_data
        self.acc_os_data = acc_os_data
        self.matching = []

        self.aseet_names_from_snipe_that_match = []
        self.asset_names_from_os_that_match = []
        self.asset_tags_match = []

        self.asset_names_from_os_that_dont_match = []
        self.non_matching = []

        self.rest_tags = []
        self.rest_names = []

    # provjera i return: lista(matching) sa brojevima os koji se nalaze u snipe
    def is_os_in_snipeit(self):
        os_in_snipe_list = []
        for acc_os_num in self.acc_os_data.acc_os_list:
            if acc_os_num in self.snipe_data.os_number:
                # print("ima ga", acc_os_num)
                self.matching.append(acc_os_num)

    # get matching in snipe and os
    def get_matcing(self):
        filename = ("usporedba_match_" + date.today().strftime("%d.%m.%Y"))

        for match in self.matching:
            self.aseet_names_from_snipe_that_match.append(self.snipe_data.asset_name[int(self.snipe_data.os_number.index(match))])

            self.asset_names_from_os_that_match.append(self.acc_os_data.acc_name[int(self.acc_os_data.acc_os_list.index(match))])

            self.asset_tags_match.append(self.snipe_data.asset_tags[int(self.snipe_data.os_number.index(match))])


    # get non-matching from os in snipe
    def get_non_matching(self):
        for non_match in self.acc_os_data.acc_os_list:
            if non_match in self.matching:
                pass
            else:
                # print(non_match)
                self.non_matching.append(non_match)
        for os_n in self.non_matching:
            self.asset_names_from_os_that_dont_match.append(self.acc_os_data.acc_name[int(self.acc_os_data.acc_os_list.index(os_n))])


    def get_rest_from_snipe(self):

        for asset_tag in self.snipe_data.dict_from_snipe_data:
            if self.snipe_data.dict_from_snipe_data[asset_tag]['os_number'] is None:
                self.rest_tags.append(asset_tag)
                self.rest_names.append(self.snipe_data.dict_from_snipe_data[asset_tag]['asset_name'])


class ExcelReport:
    def __init__(self):
        self.wb_result = Workbook()
        self.export_results_path = os.getenv("export_results_path_test")

    # upis podataka u excel
    def write_list_to_excel(self, save_name="result", start_column="A", col_name="snipe or acc", lst1=None):
        sheet_ranges_result = self.wb_result.active
        cell_n = 2
        for item in lst1:
            if item is None:
                pass
            else:
                sheet_ranges_result[start_column + "1"] = col_name
                sheet_ranges_result[start_column + str(cell_n)] = item
                cell_n += 1
                self.wb_result.save(self.export_results_path + save_name +" " + date.today().strftime("%d.%m.%Y") + ".xlsx")

    def generate_matching_xlsx(self, check):
        print("g started")
        self.write_list_to_excel(save_name="matching",start_column="A",col_name="os_num",lst1=check.matching)
        self.write_list_to_excel(save_name="matching", start_column="B", col_name="snipe_name", lst1=check.aseet_names_from_snipe_that_match)
        self.write_list_to_excel(save_name="matching",start_column="C",col_name="acc_name", lst1=check.asset_names_from_os_that_match)
        self.write_list_to_excel(save_name="matching",start_column="D",col_name="Asset_number",lst1=check.asset_tags_match)

    def generate_non_matching_xlsx(self,check):
        print("pocelo je")
        self.write_list_to_excel(save_name="non_matching",start_column="A",col_name="os_num",lst1=check.non_matching)
        self.write_list_to_excel(save_name="non_matching", start_column="B",col_name="os_name",lst1=check.asset_names_from_os_that_dont_match)


    def generate_rest_xlsx(self,check):
        self.write_list_to_excel(save_name="rest", start_column="A", col_name="asset_tages", lst1=check.rest_tags)
        self.write_list_to_excel(save_name="rest", start_column="B", col_name="name_from_snipe", lst1=check.rest_names)


"""
def test_diff():
    with open("test_files/test1.json","r") as json1:
        js1 = json.load(json1)

    with open("test_files/test2.json", "r") as json2:
        js2 = json.load(json2)

    differences = dict(DeepDiff(js1, js2))
    len_differences = len(differences["values_changed"])
    print(differences)
    # print(len_differences)
    # print(differences["values_changed"])


def main():
    global matching
    merged_raw_data_from_snipe()
    get_all_data_from_snipe()
    append_lists_from_excel()
    is_os_in_snipeit()
    create_dict_from_snipe_data()


def optional():
    # get_matcing()
    # get_non_matching()
    # test_diff()
    # get_rest_from_snipe()
    
"""

def main():
    my_snipe = Snipe()
    my_snipe.get_merged_raw_data_from_snipe()
    my_snipe.get_all_data_from_snipe()
    my_snipe.create_dict_from_snipe_data()

    my_acc = AccOsData(my_snipe)
    my_acc.append_lists_from_excel()

    my_check = Check(snipe_data=my_snipe,acc_os_data=my_acc)
    my_check.is_os_in_snipeit()
    my_check.get_matcing()
    my_check.get_non_matching()
    my_check.get_rest_from_snipe()

    my_report = ExcelReport()
    # my_report.generate_matching_xlsx(my_check)
    # my_report.generate_non_matching_xlsx(my_check)
    my_report.generate_rest_xlsx(my_check)


def my_diff():
    diff.Diff("test1", "test2").get_diff()


if __name__ == "__main__":
    # my_diff()
    main()
