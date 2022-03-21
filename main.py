from openpyxl import load_workbook
from openpyxl import Workbook

import get_os_from_snipe

wb = load_workbook(filename="test_x.xlsx")
sheet_ranges = wb["1"]
wb_result = Workbook()
sheet_ranges_result = wb_result.active

acc_os_list = []
snipe_os_list = []

#učitava podatke iz tablice u početne liste
def append_lists_from_excel():

    for cell_a in sheet_ranges["A"]:
        cell_a_value = isinstance(cell_a.value, int)
        if cell_a_value:
            if cell_a.value not in acc_os_list:
                acc_os_list.append(cell_a.value)

    for cell_c in sheet_ranges["C"]:
        cell_c_value = isinstance(cell_c.value, int)
        if cell_c_value:
            if cell_c.value not in snipe_os_list:
                snipe_os_list.append(cell_c.value)

"""
def is_os_in_snipe():
    append_lists_from_excel()
    cell_n = 2
    for acc_os in acc_os_list:

        if acc_os in snipe_os_list:
            print("ima ga ", acc_os)
            sheet_ranges_result["A1"] = "acc_os"
            sheet_ranges_result["A"+str(cell_n)] = acc_os
            print(cell_n)
            cell_n += 1
        else:
            print("nema ga ", acc_os)
"""

#provjera i return: lista sa brojevima os koji se nalaze u snipe
def is_os_in_snipeit():
    os_in_snipe_list = []
    append_lists_from_excel()

    for acc_os_num in acc_os_list:
        if acc_os_num in snipe_os_list:
            #print("ima ga", acc_os_num)
            os_in_snipe_list.append(acc_os_num)
    return  os_in_snipe_list

#upis podataka u excel
def write_to_excel(lst1=None,lst2=None):
    cell_n = 2
    #zapisuje brojeve os koji se nalaze u snipeIT-u
    for acc_os in lst1:
        sheet_ranges_result["A1"] = "acc_os"
        sheet_ranges_result["A"+str(cell_n)] = acc_os
        print(cell_n)
        cell_n += 1
        wb_result.save("result.xlsx")




if __name__ == "__main__":
    #is_os_in_snipe()

    print(type(get_os_from_snipe.merged_raw_data_from_snipe()))
    is_os_in_snipeit()
    '''
    print(acc_os_list)
    print(snipe_os_list)
    write_to_excel(lst1=is_os_in_snipeit())
    '''
