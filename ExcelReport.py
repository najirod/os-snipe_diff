import os
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import date
import sys

if "venv" in sys.path[0]:
    root_path = (sys.path[1] + "/")
else:
    root_path = (sys.path[0] + "/")


class Report:
    def __init__(self, save_path=None):
        self.wb_result = Workbook()
        # root_path = (sys.path[0] + "/")  # /Users/dpustahija1/PycharmProjects/os-snipe_diff/
        # print(root_path,"root")
        self.default_path = (root_path + (os.getenv("export_results_path_test")))
        self.save_path = save_path
        if save_path is None:
            self.export_results_path = self.default_path
        else:
            self.export_results_path = (root_path + self.save_path)
        self.sheet = self.wb_result.active

    # upis podataka u excel
    def write_list_to_excel(self, save_name="result", start_column="A", col_name="...",  lst1=None):
        sheet_ranges_result = self.sheet
        cell_n = 2
        for item in lst1:
            if item is None:
                pass
            else:
                sheet_ranges_result[start_column + "1"] = col_name
                sheet_ranges_result[start_column + str(cell_n)] = item
                cell_n += 1

                self.wb_result.save(self.export_results_path + save_name + ".xlsx")

    def write_list_to_excel_prefix(self, save_name="result", start_column="A", col_name="...", prefix_for_tag=False, lst1=None):
        sheet_ranges_result = self.sheet
        cell_n = 2
        for item in lst1:
            if item is None:
                pass
            else:
                if prefix_for_tag is True:
                    item_temp = "http://10.10.1.54/hardware/bytag?assetTag=" + item
                    full_item = '=HYPERLINK("{}", "{}")'.format(item_temp, item)
                    sheet_ranges_result[start_column + "1"] = col_name
                    sheet_ranges_result[start_column + str(cell_n)] = full_item
                    cell_n += 1

                    self.wb_result.save(self.export_results_path + save_name + ".xlsx")
                else:
                    sheet_ranges_result[start_column + "1"] = col_name
                    sheet_ranges_result[start_column + str(cell_n)] = item
                    cell_n += 1

                    self.wb_result.save(self.export_results_path + save_name + ".xlsx")

    def write_list_to_excel_with_sheets(self, save_name="result", start_column="A", col_name="...", sheet_name="Sheet", lst1=None):
        if sheet_name not in self.wb_result.sheetnames:
            self.wb_result.create_sheet(sheet_name)
            # print(self.wb_result.sheetnames)
            if "Sheet" in self.wb_result.sheetnames:
                self.wb_result.remove(self.wb_result.worksheets[0])

        sheet_ranges_result = self.wb_result[sheet_name]
        cell_n = 2
        for item in lst1:
            if item is None:
                pass
            else:
                sheet_ranges_result[start_column + "1"] = col_name
                sheet_ranges_result[start_column + str(cell_n)] = item
                cell_n += 1
                self.wb_result.save(self.export_results_path + save_name + ".xlsx")